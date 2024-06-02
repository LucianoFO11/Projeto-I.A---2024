from facenet_pytorch import MTCNN, InceptionResnetV1
import cv2
import torch
import numpy as np
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import time
import re

print('Iniciando...')  # Coloquei só pra ter certeza que tava rodando
print('Este é um programa exclusivo da Corsa Technologies,')
# Inicia o detector mtcnn
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
mtcnn = MTCNN(keep_all=True, device=device)
print(device)

# Inicia o modelo do Resnet
resnet = InceptionResnetV1(pretrained='vggface2').eval().to(device)

# Coloque aqui os rostos que o programa deve reconhecer
known_faces_folder = 'C:\\Users\\lucia\\Desktop\\rostos'

# Função para normalizar nomes de arquivos (remover parênteses e números)
def normalize_name(filename):
    name = re.sub(r'\(.*\)', '', filename)
    return name.strip()

# Carregar as imagens das pessoas conhecidas e suas respectivas "embeddings" faciais
known_faces = []
known_names = []
for filename in os.listdir(known_faces_folder):
    if filename.endswith('.jpg') or filename.endswith('.png'):
        img_path = os.path.join(known_faces_folder, filename)
        img = cv2.imread(img_path)
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        img_tensor = mtcnn(img_rgb)  # Converter para tensor
        if img_tensor is not None and len(img_tensor) > 0:
            img_tensor = img_tensor[0].to(device)  # Seleciona o primeiro tensor da lista
            embedding = resnet(img_tensor.unsqueeze(0)).detach().cpu().numpy()[0]
            known_faces.append(embedding)
            normalized_name = normalize_name(filename.split('.')[0])
            known_names.append(normalized_name)

# Dicionário para registrar a última vez que cada pessoa foi reconhecida
last_recognized = {}

# Função para registrar o reconhecimento em uma planilha Excel
def log_recognition(name):
    file_path = 'C:\\Users\\lucia\\Desktop\\reconhecimento\\reconhecimento.xlsx'
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Reconhecimentos'
        sheet.append(['Nome', 'Data', 'Hora'])
    else:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    
    now = datetime.now()
    date = now.strftime('%Y-%m-%d')
    time_str = now.strftime('%H:%M:%S')
    
    sheet.append([name, date, time_str])

    # Tentativa de salvar a planilha
    for i in range(5):  # Tentará salvar até 5 vezes
        try:
            workbook.save(file_path)
            break
        except PermissionError:
            print(f'Erro ao salvar. Tentativa {i+1}/5. Certifique-se de que a planilha esteja fechada.')
            time.sleep(1)

# Inicializar a captura de vídeo da câmera
cap = cv2.VideoCapture(0)
cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1300)
cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 700)

while True:
    ret, frame = cap.read()  # Capturar um frame da câmera
    if not ret:
        break

    # Converter o frame para RGB para detecção de faces
    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

    # Detectar rostos no frame usando MTCNN
    try:
        boxes, _ = mtcnn.detect(rgb_frame)

        # Verificar se algum rosto foi detectado
        if boxes is not None and len(boxes) > 0:
            for box in boxes:
                (x1, y1, x2, y2) = [int(coord) for coord in box]
                face = frame[y1:y2, x1:x2]
                if face.size != 0:  # Verificar se o rosto não está vazio
                    face_rgb = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)  # Converter para RGB
                    face_tensor = mtcnn(face_rgb)  # Converter para tensor
                    if face_tensor is not None and len(face_tensor) > 0:
                        face_tensor = face_tensor[0].to(device)  # Seleciona o primeiro tensor da lista
                        embedding = resnet(face_tensor.unsqueeze(0)).detach().cpu().numpy()[0]

                        # Comparar a embedding do rosto com as embeddings conhecidas
                        recognized = False
                        recognized_name = None
                        for i, known_embedding in enumerate(known_faces):
                            distance = np.linalg.norm(embedding - known_embedding)
                            if distance < 0.7:  # Defina um limite adequado para a distância
                                recognized = True
                                recognized_name = known_names[i]
                                break

                        if recognized:
                            current_time = datetime.now()
                            if (recognized_name not in last_recognized or 
                                    current_time - last_recognized[recognized_name] > timedelta(minutes=1)):
                                last_recognized[recognized_name] = current_time
                                log_recognition(recognized_name)
                                cv2.putText(frame, recognized_name, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2, cv2.LINE_AA)

                        # Desenhar retângulo ao redor do rosto detectado
                        color = (0, 255, 0) if recognized else (0, 0, 255)
                        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
    except Exception:
        pass  # Silenciosamente ignorar exceções

    # Exibir o frame processado em uma janela
    cv2.imshow('Corsa Technologies: reconhecimento facial', frame)

    # encerra o programa apertando a tecla Q
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
